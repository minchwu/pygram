# -*- coding: utf-8 -*-
# author: Minch Wu
"""文件读取.

读取个人报销信息表，提取汇总条目，输出至汇总表
"""

import os
import re
import time
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 常量定义,文件路径及底色
TMP_FILE_PATH = './output/财务汇总表.xlsx'  # 模版文件位置
INPUT_FILE_PATH = './input/'  # 数据文件位置
LOG_FILE_PATH = './fileread.log'  # 日志文件位置
OBJ_FILE_NAME = './output/' + '报销登记' + '+' + re.sub(
    '-', '', str(datetime.date.today())) + '.xlsx'  # 目标文件名
SIGN_COLOR = ['7FFFD4', '87CEFF']  # 填充底色

wb_tmp = load_workbook(TMP_FILE_PATH)
wb_tmp_sheetnames = wb_tmp.sheetnames
ws_tmp_obj = wb_tmp[wb_tmp_sheetnames[0]]  # 默认第一个sheet为数据记录表
ws_tmp_class = wb_tmp[wb_tmp_sheetnames[1]]  # 默认第二个sheet为分类明细参考表

# 建立分类索引，报销业务类别
''' 分类明细格式
1.|编号|分类|一级分类|分类说明|
2.|0  |1   |2      |3     |
n.|...
'''
data_class = {}
for row in ws_tmp_class.iter_rows(min_row=2):
    data_class[row[1].value] = row[2].value

# 报销人文件循环，默认.xlsx文件格式
input_file_list = os.listdir(INPUT_FILE_PATH)
for each_file in input_file_list:
    if os.path.splitext(each_file)[1] != '.xlsx':
        input_file_list.remove(each_file)


# 日期格式转换
def date_format(date: datetime.datetime):
    """date_format."""
    y, m, d = date.year, date.month, date.day
    return (datetime.date(y, m, d))  # Y-m-d 格式


''' 汇总文件格式
1.|业务日期|报销日期|说明|类别/对方账户|对方|费用归属|报销人|项目|部门|金额|month|大类|有无发票|电子发票号|备注（多报）|
2.|0      |1      |2  |3          |4  |5      |6    |7  |8  |9   |10   |11 |12     |13      |14        |
n.|...
'''
''' 数据文件格式
1.|                          个人差旅报销                                |
2.|部门   |xxx  |报销人     |xxx            |报销总金额  |xxx    |银行及账号|
3.|项目   |xxx  |报销提交日期|xxx(2019/01/01)|票据总数    |       |同银行账户|
4.|                                                                    |
5.|                                                                    |
6.|日期   |说明  |  类别     |参与人员        |报销金额    |票据数  |电子发票号|
7.|xxx   |xxx   |xxx       |xxx            |xxx        |xxx    |xxx     |
...
n.|                   合计                  |xxx        |xxx    |xxx     |
'''

log = open(LOG_FILE_PATH, 'a')
# 文件读取计数
file_count = 0
for each_file in input_file_list:
    print("{0:{2}<25}{1}".format(each_file, "信息读取开始>>>>>>>", chr(12288)))
    # log.writelines('{0:{2}<25}{1:{2}>30}{3}'.format(
    #     each_file, '<<time:'
    #     + str(datetime.datetime.now())
    #     + '>>', chr(12288), '\n'))
    # log.writelines("{0:{1}<25}{2}".format(
    #     "信息读取开始>>>>>>>", chr(12288), '\n'))
    log.writelines('{0}{1}{2}{3}'.format(
        '<<time: ' + str(datetime.datetime.now()) + '>>', '\n', each_file,
        '\n'))
    log.writelines('{0}{1}'.format('信息读取开始......', '\n'))

    wb_data = load_workbook(INPUT_FILE_PATH + each_file)
    wb_data_sheetnames = wb_data.sheetnames

    color = SIGN_COLOR[file_count % 2]

    for each_sheet in wb_data_sheetnames:
        if each_sheet != u'分类明细参考':
            ws_data = wb_data[each_sheet]

            # 报销人信息提取
            department = ws_data.cell(row=2, column=2).value  # 部门
            project = ws_data.cell(row=3, column=2).value  # 项目
            person = ws_data.cell(row=2, column=4).value  # 报销人
            date_sub = date_format(ws_data.cell(row=3,
                                                column=4).value)  # 报销提交日期

            # 报销业务信息提取
            rec_count = 0
            for row in ws_data.iter_rows(min_row=7):  # 默认第七行开始为具体业务记录
                try:
                    rec_count += 1
                    error_loc = 0  # 0
                    time_cache = date_format(row[0].value)
                    error_loc += 1  # 1
                    time_struct = time.strptime(str(time_cache),
                                                '%Y-%m-%d')  # 日期格式化数组
                    error_loc += 1  # 2
                    time_str = time.strftime('%Y/%m/%d',
                                             time_struct)  # 业务日期(Y/m/d)
                    error_loc += 1  # 3
                    time_ym = time.strftime('%Y-%m', time_struct)  # month(y-m)

                    # 单条数据汇总
                    error_loc += 1  # 4
                    data_sum = [''] * 15
                    error_loc += 1  # 5
                    data_sum[0] = time_str
                    error_loc += 1  # 6
                    data_sum[1] = date_sub
                    error_loc += 1  # 7
                    data_sum[2] = row[1].value + ';' + row[3].value
                    error_loc += 1  # 8
                    data_sum[3] = row[2].value
                    error_loc += 1  # 9
                    data_sum[6] = person
                    error_loc += 1  # 10
                    data_sum[7] = project
                    error_loc += 1  # 11
                    data_sum[8] = department
                    error_loc += 1  # 12
                    data_sum[9] = float(('%.f' % row[4].value))
                    error_loc += 1  # 13
                    data_sum[10] = time_ym
                    error_loc += 1  # 14
                    data_sum[11] = data_class[row[2].value]
                    error_loc += 1  # 16
                    data_sum[12] = ['无', '有'][int(bool(row[5].value))]
                    error_loc += 1  # 17
                    data_sum[13] = row[6].value

                    error_loc += 1  # 18
                    ws_tmp_obj.append(data_sum)

                    # 金额单元格定位并填充底色
                    error_loc += 1  # 19
                    nrows = ws_tmp_obj.max_row
                    error_loc += 1  # 20
                    ws_tmp_obj.cell(row=nrows, column=10).fill = PatternFill(
                        fill_type='solid', fgColor=color)

                except Exception as error:
                    print(error)
                    log.writelines(
                        "Error: <sheet:{0}, 第{1}行记录, {2}, ERROR_LOC = {3}>{4}".
                        format(each_sheet, rec_count + 6, error, error_loc,
                               '\n'))
                    break

    file_count += 1
    print(">>>>>>>信息读取结束\n")
    # log.writelines("{0:{1}<25}{2}".format(
    #     ">>>>>>>信息读取结束", chr(12288), '\n\n'))
    log.writelines('{0}{1}'.format('信息读取结束', '\n\n'))

    # # 日志文件记录
    # with open(LOG_FILE_PATH, 'a') as log:
    #     log.writelines('{0:{2}<30}{1:{2}>40}{3}'.format(
    #         each_file, '<<time:'
    #         + str(datetime.datetime.now())
    #         + '>>', chr(12288), '\n'))

log.close()
wb_tmp.save(OBJ_FILE_NAME)
print("you can check your fileread log file in details")
