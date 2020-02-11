# -*- coding: utf-8 -*-
# author: Minch Wu
"""xiujia.

读取每个月的请假信息，提取信息至年度休假表，并做简单统计
1. 工作日的请假情况，年假，病假，事假，并判断是一天还是半天
    以4小时为分割点，对请假类型进行底色填充，浅色
2. 每个人的年假天数是一定的，统计剩余天数
3. 特别地，存在补假情况，一天分成时间段给出，申请日期滞后请假日期
4. 请假单位为半天，存在半天以下的情况，添加注释
"""

import os
import re
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

# 常量定义
if True:
    # 定义块
    TMP_PATH = "./output/2018年休假记录及统计表.xlsx"  # 模板文件
    OBJ_PATH = "./output/2018年休假记录及汇总表@{}.xlsx".format(
        datetime.date.today())  # 模板文件另存文件名
    INPUT_PATH = "./input/"  # 数据文件路径
    LOG_FILE_PATH = "./fileread.log"
    COMMENT_AUTHOR = "柳秀秀"  # 注释人
    SIGN_COLOR = {
        '病假一天': '87CEFF',
        '病假半天': 'F4A460',
        '事假一天': 'C4C4C4',
        '事假半天': 'EEB422',
        '年假一天': '4876FF',
        '年假半天': '3CB371',
        '婚假一天': '3CB231',
        '婚假半天': '34FEAC',
        '产假一天': 'AE23DF',
        '产假半天': 'DDCCEE',
    }  # 填充色号
    DAY_COUNT = {
        '病假一天': ['病假', 1],
        '病假半天': ['病假', 0.5],
        '事假一天': ['事假', 1],
        '事假半天': ['事假', 0.5],
        '年假一天': ['年假', 1],
        '年假半天': ['年假', 0.5],
        '婚假一天': ['婚假', 1],
        '婚假半天': ['婚假', 0.5],
        '产假一天': ['产假', 1],
        '产假半天': ['产假', 0.5],
        None: [None, 0],
    }

# 打开模板文件
wb_tmp = load_workbook(TMP_PATH)
wb_tmp_sheet_names = wb_tmp.sheetnames
ws_rec = wb_tmp[wb_tmp_sheet_names[0]]  # 数据记录表
ws_sum = wb_tmp[wb_tmp_sheet_names[1]]  # 数据汇总表
# ws_ref = wb_tmp[wb_tmp_sheet_names[2]]    # 参考信息表（2019）

# 建立数据文件列表，默认.xlsx数据源
input_file_list = os.listdir(INPUT_PATH)
for each_file in input_file_list:
    if os.path.splitext(each_file)[1] != '.xlsx':
        input_file_list.remove(each_file)

# 记录表的索引字典
ws_rec_row_index, i = {}, 4
ws_rec_col_index, j = {}, 3

# 请假信息统计
ws_rec_id_sum = {}  # FIXME: 对于连续请假日期中的休息日应当不予记录，采用读表的方法统计休假信息

# 行索引
for row in ws_rec.iter_rows(min_row=4):
    try:  # 分割日期键值和统计键值
        date_cache = row[0].value
        y = date_cache.year
        m = date_cache.month
        d = date_cache.day
        key = str(datetime.date(y, m, d))
        ws_rec_row_index[key] = i
        i += 1
    except:
        key = row[0].value
        ws_rec_row_index[key] = i
        i += 1

# 列索引
for col in ws_rec.iter_cols(min_col=3):
    ws_rec_col_index[str(col[1].value)] = j
    a, b, c, d, e = (
        ws_rec.cell(row=ws_rec_row_index['病假'], column=j).value,
        ws_rec.cell(row=ws_rec_row_index['事假'], column=j).value,
        ws_rec.cell(row=ws_rec_row_index['年假'], column=j).value,
        ws_rec.cell(row=ws_rec_row_index['婚假'], column=j).value,
        ws_rec.cell(row=ws_rec_row_index['产假'], column=j).value,
    )
    ws_rec_id_sum[str(col[1].value)] = {
        '病假': [a, 0][int(a == None)],
        '事假': [b, 0][int(b == None)],
        '年假': [c, 0][int(c == None)],
        '婚假': [d, 0][int(d == None)],
        '产假': [e, 0][int(e == None)],
        None: 0
    }
    j += 1

# print(ws_rec_row_index)
# print(ws_rec_col_index)
# print(ws_rec_id_sum)
""" 数据源格式
|0  |1  |2      |3      |4      |5         |6      |7       |8      ...  |18      |
工号|姓名|部门/团队|申请日期|请假类型|请假时间类型|请假日期|请假时间段|请假原因|。。。|最终状态 |
"""


# 数据处理函数
def read_data_to_obj(wb_data: Workbook):
    ''' 读取单月请假信息至目标文件 '''
    wb_data_sheet_names = wb_data.sheetnames
    ws_data = wb_data[wb_data_sheet_names[0]]  # 默认第一个sheet为数据源

    # 请假条目计数器
    # ask_count = 0

    for row in ws_data.iter_rows(min_row=2):

        job_id = row[0].value.split('_')[1]  # 工号(QXZN_xxxx)
        name = row[1].value  # 姓名
        application_date = row[3].value  # 申请日期(Y-m-d h:m)
        ask_class = row[4].value  # 请假类型(年假/病假/事假)
        ask_date_class = row[5].value  # 请假时间类型(按时间段/按天)
        ask_reason = row[8].value  # 请假原因
        yorn = row[18].value  # 最终状态

        if yorn == '批准':
            read_log = "QXZN_{0}({1})于<{2}>申请<{3}>, 请假时间类型：<{4}>, 请假原因：<{5}>\n信息读取中......".format(
                job_id, name, application_date, ask_class, ask_date_class,
                ask_reason)
            print(read_log)

            coln = ws_rec_col_index[job_id]  # 定位单元格列(工号)

            # 对请假时间类型为"按时间段"的数据进行时间段处理
            if ask_date_class == u'按时间段':
                # try:
                ask_date = row[6].value  # 请假日期(Y-m-d)
                rown = ws_rec_row_index[ask_date]  # 定位输出单元格(请假日期)

                # 请假时间段 09:00 ~ 12:00, 13:00 ~ 18:00
                ask_date_time = row[7].value
                cache_str = re.sub('\s', '', ask_date_time)  # 替换时间段中的空格
                cache_list = [(lambda x: int(x))(x)
                              for x in re.split(':|~|,', cache_str)
                              ]  # 分割时间段字符串，得到时间节点列表
                ask_date_time_len = sum([
                    (cache_list[4 * i + 2] + cache_list[4 * i + 3] / 60) -
                    (cache_list[4 * i] + cache_list[4 * i + 1] / 60)
                    for i in range(len(cache_list) // 4)
                ])  # 计算时间段总时长，判断请假单位（天/半天）

                # 时间段 大于四小时 判断为一天，其余按半天算，注释请假时长
                if ask_date_time_len > 4:
                    # ws_rec_id_sum[job_id][ask_class] += 1  # job_id 相应请假类型统计加1
                    date_class = u'一天'
                    value = ask_class + date_class  # 填充值
                    ws_rec.cell(row=rown, column=coln, value=value)
                    ws_rec.cell(row=rown, column=coln).fill = PatternFill(
                        fill_type='solid', fgColor=SIGN_COLOR[value])  # 底色填充

                    if str(application_date) > str(
                            ask_date):  # 申请日期落后于请假日期定义为补假
                        # TODO: 补假验证，补假情况仅出现在时间段类型中，（排除特殊性请假时间段为一天的情况，补假的申请日期落后于请假日期）
                        comment = Comment(
                            u"补假{}".format(
                                ["{}小时".format(ask_date_time_len),
                                 "一天"][int(ask_date_time_len == 8)]),
                            author=COMMENT_AUTHOR)  # 补假注释，补假8小时定义为补假一天，其余按小时注释
                        ws_rec.cell(row=rown, column=coln).comment = comment
                    else:
                        # comment = Comment(u"请假{}小时".format(ask_date_time_len), author=COMMENT_AUTHOR)
                        comment = Comment(
                            u"请假{}".format(
                                ["{}小时".format(ask_date_time_len),
                                 "一天"][int(ask_date_time_len == 8)]),
                            author=COMMENT_AUTHOR)  # 请假注释，请假8小时定义为请假一天，其余按小时注释
                        ws_rec.cell(row=rown, column=coln).comment = comment
                else:
                    # ws_rec_id_sum[job_id][ask_class] += 0.5    # job_id 相应请假类型统计加0.5
                    date_class = u'半天'
                    value = ask_class + date_class  # 填充值
                    ws_rec.cell(row=rown, column=coln, value=value)
                    ws_rec.cell(row=rown, column=coln).fill = PatternFill(
                        fill_type='solid', fgColor=SIGN_COLOR[value])

                    if str(application_date) > str(
                            ask_date):  # 申请日期落后于请假日期定义为补假
                        comment = Comment(u"补假{}".format(
                            ["{}小时".format(ask_date_time_len),
                             "半天"][int(ask_date_time_len == 4)]),
                                          author=COMMENT_AUTHOR
                                          )  # 补假注释， 补假4小时定义为补假半天，其余按小时注释
                        ws_rec.cell(row=rown, column=coln).comment = comment
                    else:
                        comment = Comment(u"请假{}".format(
                            ["{}小时".format(ask_date_time_len),
                             "半天"][int(ask_date_time_len == 4)]),
                                          author=COMMENT_AUTHOR
                                          )  # 补假注释， 补假4小时定义为补假半天，其余按小时注释
                        ws_rec.cell(row=rown, column=coln).comment = comment

            # except Exception as error:
            # print(error)

            elif ask_date_class == u'按天':
                # try:
                # 请假日期 2019-01-01 - 2019-01-01 //  2019-01-01 - 2019-01-02
                ask_date = row[6].value  # 请假日期
                cache_list = ask_date.split()  # 删除空格
                cache_list.remove('-')  # 删除日期分隔符
                start, end = cache_list
                day_start, day_end = datetime.datetime.strptime(
                    start, '%Y-%m-%d'), datetime.datetime.strptime(
                        end, '%Y-%m-%d')  # 构造日期类型起止时间节点

                date_class = u'一天'
                value = ask_class + date_class
                while day_start <= day_end:
                    try:  # 此处可能捕捉到连续请假时间中的休息日，统计表中没有对应时间，跳过该日期，循环继续
                        # ws_rec_id_sum[job_id][ask_class] += 1
                        rown = ws_rec_row_index[datetime.datetime.strftime(
                            day_start, '%Y-%m-%d')]
                        ws_rec.cell(row=rown, column=coln, value=value)
                        ws_rec.cell(row=rown, column=coln).fill = PatternFill(
                            fill_type='solid', fgColor=SIGN_COLOR[value])
                        day_start += datetime.timedelta(days=1)
                    except Exception as error:
                        day_start += datetime.timedelta(days=1)
                        print(error, "正常休息日，不需要请假")
                        continue

            # except Exception as error:
            # print(error)
            # pass

            print("信息读取完成\n")


# 对于数据列表中的文件循环处理
for each_file in input_file_list:
    print(each_file, "信息读取中>>>>>>>\n")
    read_data_to_obj(load_workbook(INPUT_PATH + each_file))
    print(each_file, ">>>>>>>信息读取完成\n\n")

# 请假天数统计
for col in ws_rec.iter_cols(min_col=3):
    key = str(col[1].value)
    for row in col[3:-5]:
        ws_rec_id_sum[key][DAY_COUNT[row.value][0]] += DAY_COUNT[row.value][1]
    # for row in col[-3:]:
# print(ws_rec_id_sum)

for each_id in ws_rec_col_index:
    ws_rec.cell(row=ws_rec_row_index['病假'],
                column=ws_rec_col_index[each_id],
                value=ws_rec_id_sum[each_id]['病假'])
    ws_rec.cell(row=ws_rec_row_index['事假'],
                column=ws_rec_col_index[each_id],
                value=ws_rec_id_sum[each_id]['事假'])
    ws_rec.cell(row=ws_rec_row_index['年假'],
                column=ws_rec_col_index[each_id],
                value=ws_rec_id_sum[each_id]['年假'])
    ws_rec.cell(row=ws_rec_row_index['婚假'],
                column=ws_rec_col_index[each_id],
                value=ws_rec_id_sum[each_id]['婚假'])
    ws_rec.cell(row=ws_rec_row_index['产假'],
                column=ws_rec_col_index[each_id],
                value=ws_rec_id_sum[each_id]['产假'])

# 模板文件另存为处理文件
wb_tmp.save(OBJ_PATH)
print("数据处理成功!!!")

# if __name__ == '__main__':
#     pass
# else:
#     pass
